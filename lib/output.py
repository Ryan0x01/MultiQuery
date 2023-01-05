import os
import re

import pymysql
from prettytable import PrettyTable
import openpyxl as xl

from api.database import Database


class Output:
    def __init__(self, data_list, out_type, file_name=''):
        self.title_list = ['协议', 'IP地址', '端口', '域名', '组件/版本', '站点标题', '备案号']
        self.data_list = data_list
        self.file_name = file_name
        if out_type == 'print':
            self.printTable()
        elif out_type == 'excel':
            self.saveToExcel()
        elif out_type == 'db':
            self.db = Database()
            self.db.initDatabase()
            self.db.createTable()
            self.saveToDatabase()

    def printTable(self):
        """
        以表格的形式对数据进行整理输出
        """
        table = PrettyTable(self.title_list)
        row = []
        for data in self.data_list:
            for i in range(7):
                if (i == 4 or i == 5) and len(data[i]) > 20:
                    data[i] = data[i][0:20]
                    data[i] += "..."
                row.append(data[i].strip())
            table.add_row(row)
            row.clear()
        print(table)

    def saveToExcel(self):
        # 定义Excel非法字符
        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
        # 如果文件保存路径不存在则创建
        if not os.path.exists("output"):
            os.mkdir("output")
        file_path = "output/" + self.file_name + ".xlsx"
        # 将数据写入excel，如果不存在则创建，存在则替换
        col_width = {'A': 10, 'B': 20, 'C': 10, 'D': 30, 'E': 25, 'F': 40, 'G': 20}
        wb = xl.Workbook()
        ws = wb.active
        ws.title = self.file_name
        for i in range(0, len(self.title_list)):
            ws.cell(1, i + 1).value = self.title_list[i]
        for k, v in col_width.items():
            ws.column_dimensions[k].width = v
        ws.freeze_panes = 'A2'
        # 插入数据，将端口一列的单元格类型设置为数字
        for row in range(2, len(self.data_list)):
            for col in range(1, len(self.title_list)):
                # 清理Excel非法字符
                value = ILLEGAL_CHARACTERS_RE.sub(r'', self.data_list[row-2][col-1])
                # 写入数据
                ws.cell(row, col).value = value
                # 设置端口一列为整数
                if col == 3:
                    ws.cell(row, 3).data_type = "int"
        try:
            wb.save(file_path)
            print('查询结果已保存在{}中'.format(file_path))
        except PermissionError:
            print("目标写入文件已打开，无法写入数据，请关闭文件后重试！")
            return -1

    def saveToDatabase(self):
        convertor = pymysql.converters
        self.printTable()
        for data in self.data_list:
            # 对单引号和双引号进行转义
            data[4] = convertor.escape_string(str(data[4]))
            data[5] = convertor.escape_string(str((data[5])[0:255]))
            self.db.insertIntoTable(data)
        print("数据已成功写入数据库!")
